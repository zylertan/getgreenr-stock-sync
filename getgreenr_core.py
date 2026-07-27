"""
getgreenr_core.py — Apply the stock rules to the GetGreenr Seller Stock column.

Four inputs (matching the reviewed workflow):
  1. Masterlist Excel        -> live Used Available Qty per Stock Type ID
  2. GetGreenr bulk stock     -> the file to update (Seller Stock column)
  3. SKU Registry workbook    -> Locked mappings + Match Review reviewer decisions
                                 (+ optional Accessories sheet)
  4. New Masterlist SKUs      -> informational (surfaced in the report)

Rules applied to Seller Stock, keyed by the GetGreenr SKU ID:
  - Locked Matches      -> SUM of live Used Available Qty across the linked Masterlist IDs
  - Accessories         -> 10
  - Match Review        -> reviewer decision: "Set to 0"/"Skip / Delist" -> 0,
                           otherwise (Link manually / Keep as-is / blank) -> leave unchanged
  - Anything else       -> leave unchanged

The original GetGreenr file format is preserved (CSV round-trip, or in-place
openpyxl edit for .xlsx) so the output is ready to re-upload.
"""
from __future__ import annotations

import io
import re
import pandas as pd
from openpyxl import load_workbook

import matcher as M

# Priority order for locating the stock column in the GetGreenr file
STOCK_CANDS = ["seller stock", "sellerstock", "on_hand", "on hand", "onhand",
               "available", "stock", "quantity", "qty"]
SKU_CANDS = ["sku", "sku id", "getgreenr sku id", "seller sku", "listing sku"]

# Reviewer-decision values that force stock to 0
ZERO_DECISIONS = {"set to 0", "set 0", "skip / delist", "skip/delist", "delist", "skip"}


def _norm(s):
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]", " ", str(s).strip().lower())).strip()


def find_col(cols, cands):
    nmap = {c: _norm(c) for c in cols}
    # exact match first
    for cand in cands:
        for c, n in nmap.items():
            if n and n == cand:
                return c
    # substring match (guard against empty/very short headers matching everything)
    for cand in cands:
        for c, n in nmap.items():
            if not n:
                continue
            if cand in n or (len(n) >= 4 and n in cand):
                return c
    return None


def to_int(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return 0
    m = re.search(r"-?\d+(?:\.\d+)?", str(v).replace(",", ""))
    return max(int(round(float(m.group()))), 0) if m else 0


def nid(v):
    """Normalise a Stock Type ID for matching (drops trailing '.0')."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return re.sub(r"\.0$", "", str(v).strip())


def _blank(v):
    return v is None or (isinstance(v, float) and pd.isna(v)) or str(v).strip() == "" or str(v).strip().lower() == "nan"


# --------------------------------------------------------------------------- #
# Registry reading
# --------------------------------------------------------------------------- #
def read_registry(registry_file):
    """Return (locked_map, review_map, accessory_set).

    locked_map : sku -> {"ids": [masterlist ids], "target": registry Target Qty}
    review_map : sku -> reviewer decision (lower-case)
    accessory_set : set of sku
    """
    sheets = pd.read_excel(registry_file, sheet_name=None, dtype=object)

    def sheet_like(*subs):
        for name, df in sheets.items():
            n = _norm(name)
            if any(s in n for s in subs):
                return df
        return None

    locked_map, review_map, accessory_set = {}, {}, set()

    lk = sheet_like("locked")
    if lk is not None:
        sku_c = find_col(lk.columns, ["getgreenr sku id", "sku id", "sku", "gg_sku"])
        id_c = find_col(lk.columns, ["locked masterlist id", "masterlist id", "linked"])
        tgt_c = find_col(lk.columns, ["target qty", "ml available qty", "summed"])
        for _, r in lk.iterrows():
            sku = r.get(sku_c)
            if sku is None or (isinstance(sku, float) and pd.isna(sku)):
                continue
            ids = []
            if id_c is not None and not (isinstance(r.get(id_c), float) and pd.isna(r.get(id_c))):
                ids = [nid(x) for x in re.split(r"[,\s]+", str(r.get(id_c))) if x.strip()]
            locked_map[str(sku).strip()] = {
                "ids": ids, "target": to_int(r.get(tgt_c)) if tgt_c else 0}

    mr = sheet_like("match review")
    if mr is not None:
        sku_c = find_col(mr.columns, ["getgreenr sku id", "sku id", "sku", "gg_sku"])
        dec_c = find_col(mr.columns, ["reviewer decision", "decision"])
        for _, r in mr.iterrows():
            sku = r.get(sku_c)
            if sku is None or (isinstance(sku, float) and pd.isna(sku)):
                continue
            dec = r.get(dec_c) if dec_c else None
            review_map[str(sku).strip()] = _norm(dec) if dec is not None and not (
                isinstance(dec, float) and pd.isna(dec)) else ""

    acc = sheet_like("accessor")
    if acc is not None:
        sku_c = find_col(acc.columns, ["getgreenr sku id", "sku id", "sku", "gg_sku"])
        if sku_c is not None:
            for _, r in acc.iterrows():
                sku = r.get(sku_c)
                if sku is not None and not (isinstance(sku, float) and pd.isna(sku)):
                    accessory_set.add(str(sku).strip())

    # New Masterlist SKUs marked "Linked" -> fold into the locked mapping so the
    # linked Masterlist ID's Used qty flows to the chosen GetGreenr listing.
    nm = sheet_like("new masterlist")
    if nm is not None:
        link_c = find_col(nm.columns, ["link to getgreenr sku id", "link to sku", "getgreenr sku id"])
        mlid_c = find_col(nm.columns, ["masterlist stock type id", "stock type id", "masterlist id"])
        dec_c = find_col(nm.columns, ["reviewer decision", "decision"])
        for _, r in nm.iterrows():
            if dec_c and _norm(r.get(dec_c)) == "linked" and link_c and not _blank(r.get(link_c)):
                link = str(r.get(link_c)).strip()
                mlid = nid(r.get(mlid_c)) if mlid_c else ""
                entry = locked_map.setdefault(link, {"ids": [], "target": 0})
                if mlid and mlid not in entry["ids"]:
                    entry["ids"].append(mlid)

    return locked_map, review_map, accessory_set


def new_masterlist_status(registry_file):
    """(total, reviewed, missing) rows on the New Masterlist SKUs sheet.

    A row counts as reviewed when its Reviewer Decision is non-blank.
    """
    sheets = pd.read_excel(registry_file, sheet_name=None, dtype=object)
    df = None
    for name, d in sheets.items():
        if "new masterlist" in _norm(name):
            df = d
            break
    if df is None:
        return (0, 0, 0)
    df = df.dropna(how="all")
    total = len(df)
    dec_c = find_col(df.columns, ["reviewer decision", "decision"])
    if dec_c is None:
        return (total, 0, total)
    reviewed = int(df[dec_c].apply(lambda v: not _blank(v)).sum())
    return (total, reviewed, total - reviewed)


# --------------------------------------------------------------------------- #
# Masterlist live Used qty
# --------------------------------------------------------------------------- #
def masterlist_used_qty(ml_file):
    ml = M.load_masterlist(ml_file)
    ml = ml[ml["Category"].astype(str).str.strip().str.lower() == "used"]
    qty = {}
    for _, r in ml.iterrows():
        qty[nid(r["StockTypeID"])] = qty.get(nid(r["StockTypeID"]), 0) + int(r["qty"])
    return qty


# --------------------------------------------------------------------------- #
# Main apply
# --------------------------------------------------------------------------- #
ACCESSORY_STOCK = 10


def run(getgreenr_file, masterlist_file, registry_file,
        new_skus_file=None, stock_col=None, sku_col=None, filename=""):
    errors = []

    def err(level, cat, detail, **x):
        errors.append({"level": level, "category": cat, "detail": detail, **x})

    is_csv = str(filename or getattr(getgreenr_file, "name", "")).lower().endswith(".csv")
    raw_bytes = getgreenr_file.read() if hasattr(getgreenr_file, "read") else open(getgreenr_file, "rb").read()

    if is_csv:
        gg = pd.read_csv(io.BytesIO(raw_bytes), dtype=object)
    else:
        gg = pd.read_excel(io.BytesIO(raw_bytes), dtype=object)

    sku_col = sku_col or find_col(gg.columns, SKU_CANDS)
    stock_col = stock_col or find_col(gg.columns, STOCK_CANDS)
    if sku_col is None:
        err("ERROR", "getgreenr", "Could not find the SKU ID column in the GetGreenr file.")
    if stock_col is None:
        err("ERROR", "getgreenr", "Could not find the Seller Stock column in the GetGreenr file.")

    locked_map, review_map, accessory_set = read_registry(registry_file)
    used_qty = masterlist_used_qty(masterlist_file) if masterlist_file is not None else {}

    counts = {"locked": 0, "accessories": 0, "review_zero": 0,
              "review_unchanged": 0, "unmatched_unchanged": 0}
    preview = []

    if sku_col is not None and stock_col is not None:
        new_vals = []
        for _, row in gg.iterrows():
            sku = str(row.get(sku_col)).strip()
            old = to_int(row.get(stock_col))
            rule, new = "unchanged", old

            if sku in locked_map:
                ids = locked_map[sku]["ids"]
                if used_qty and ids:
                    total, missing = 0, []
                    for i in ids:
                        if i in used_qty:
                            total += used_qty[i]
                        else:
                            missing.append(i)
                    if missing:
                        err("WARNING", "locked",
                            f"Masterlist ID(s) not found (Used) — treated as 0: {', '.join(missing)}",
                            sku=sku)
                    new = total
                else:
                    new = locked_map[sku]["target"]  # fall back to registry Target Qty
                rule = "locked"
                counts["locked"] += 1
            elif sku in accessory_set:
                new, rule = ACCESSORY_STOCK, "accessory"
                counts["accessories"] += 1
            elif sku in review_map:
                if review_map[sku] in ZERO_DECISIONS:
                    new, rule = 0, "review->0"
                    counts["review_zero"] += 1
                else:
                    rule = "review-unchanged"
                    counts["review_unchanged"] += 1
            else:
                counts["unmatched_unchanged"] += 1

            new_vals.append(new)
            preview.append({"SKU": sku, "Rule": rule, "Old": old, "New": new,
                            "Changed": new != old})
        gg[stock_col] = new_vals

    # New Masterlist SKUs (informational)
    new_ct = 0
    if new_skus_file is not None:
        try:
            nsk = pd.read_excel(new_skus_file, dtype=object)
            new_ct = len(nsk.dropna(how="all"))
        except Exception as e:  # noqa: BLE001
            err("WARNING", "new_sku", f"Could not read New Masterlist SKUs: {e}")

    summary = {
        "GetGreenr listings": len(gg),
        "Locked (synced to Used qty)": counts["locked"],
        "Accessories set to 10": counts["accessories"],
        "Match Review set to 0": counts["review_zero"],
        "Match Review left unchanged": counts["review_unchanged"],
        "Unmatched left unchanged": counts["unmatched_unchanged"],
        "Stock column updated": stock_col,
        "SKU column used": sku_col,
        "New Masterlist SKUs listed": new_ct,
        "Warnings": sum(1 for e in errors if e["level"] == "WARNING"),
        "Errors": sum(1 for e in errors if e["level"] == "ERROR"),
    }

    # ---- Build output bytes preserving format ----
    out_bytes, out_name = None, None
    if sku_col is not None and stock_col is not None:
        if is_csv:
            buf = io.StringIO()
            gg.to_csv(buf, index=False)
            out_bytes = buf.getvalue().encode("utf-8")
            out_name = "GetGreenr_bulk_stock_UPDATED.csv"
        else:
            wb = load_workbook(io.BytesIO(raw_bytes))
            ws = wb.active
            headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
            scol = headers.get(stock_col)
            kcol = headers.get(sku_col)
            if scol and kcol:
                lookup = {str(r[sku_col]).strip(): v for r, v in
                          zip(gg.to_dict("records"), gg[stock_col])}
                for rr in range(2, ws.max_row + 1):
                    k = str(ws.cell(rr, kcol).value).strip()
                    if k in lookup:
                        ws.cell(rr, scol).value = int(lookup[k])
            bio = io.BytesIO(); wb.save(bio); out_bytes = bio.getvalue()
            out_name = "GetGreenr_bulk_stock_UPDATED.xlsx"

    return {
        "updated_df": gg, "summary": summary, "errors": errors,
        "preview": pd.DataFrame(preview), "out_bytes": out_bytes, "out_name": out_name,
        "stock_col": stock_col, "sku_col": sku_col,
    }
