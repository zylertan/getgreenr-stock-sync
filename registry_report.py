"""
registry_report.py — Write the GetGreenr SKU registry as a SellUp-styled workbook.

Replicates the SellUp_Match_Review design:
  - navy (#1F3864) '#' column header w/ white text
  - orange (#F4B183) header band for the source (GetGreenr) columns
  - yellow (#FFD966) header band for the Masterlist / match columns
  - frozen header row (A2), autofilter, sensible column widths
  - a titled Summary sheet with an offset Category/Count table
Sheet set mirrors SellUp, with SellUp -> GetGreenr renaming.
"""
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import matcher as M

NAVY = "1F3864"
ORANGE = "F4B183"
YELLOW = "FFD966"
WHITE = "FFFFFF"
GREY_RULE = "D9D9D9"

OUT = "/sessions/dreamy-hopeful-mayer/mnt/outputs/GetGreenr_Match_Review.xlsx"

thin = Side(style="thin", color=GREY_RULE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# --------------------------------------------------------------------------- #
# Display-field helpers
# --------------------------------------------------------------------------- #
def disp_model(sku):
    """First segment of the GetGreenr sku = 'Brand + Model' (e.g. 'Google Pixel 10 Pro XL')."""
    return str(sku).split(" - ")[0].strip()


def disp_colour(sku):
    """Last non-spec segment of the sku, title-cased."""
    segs = [s.strip() for s in str(sku).replace("–", "-").split(" - ") if s.strip()]
    for seg in reversed(segs[1:]):
        low = seg.lower()
        if re.search(r"\d+\s*(gb|tb)", low) or "sim" in low or "ram" in low or "wifi" in low:
            continue
        return seg
    return ""


def connectivity(sku):
    low = str(sku).lower()
    if "wifi" in low or "wi-fi" in low:
        return "WiFi"
    if "cellular" in low or "cell" in low:
        return "Cellular"
    return ""


def category(sku):
    low = str(sku).lower()
    if any(k in low for k in ["ipad", "magicpad", " pad", "pad "]):
        return "Tablets"
    if "watch" in low:
        return "Watches"
    if "airpod" in low or "buds" in low:
        return "Audio"
    return "Smartphones"


# --------------------------------------------------------------------------- #
# Styling
# --------------------------------------------------------------------------- #
def write_sheet(wb, title, headers, band_colors, rows, widths, dropdowns=None):
    ws = wb.create_sheet(title)
    for c, (h, band) in enumerate(zip(headers, band_colors), start=1):
        cell = ws.cell(1, c, h)
        cell.fill = PatternFill("solid", fgColor=band)
        cell.font = Font(bold=True, color=(WHITE if band == NAVY else "000000"), size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(r, c, val)
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = BORDER
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    ws.row_dimensions[1].height = 30

    # Dropdowns: list of (col_index, [options]) -> Excel data-validation list
    for col_idx, options in (dropdowns or []):
        col = get_column_letter(col_idx)
        dv = DataValidation(type="list",
                            formula1='"' + ",".join(options) + '"',
                            allow_blank=True, showDropDown=False)
        dv.error = "Please choose from the list"
        dv.prompt = "Select a decision"
        # apply to a generous range so new rows also get the dropdown
        last = max(len(rows) + 1, 2)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{last}")
    return ws


def summary_sheet(wb, counts):
    ws = wb.create_sheet("Summary", 0)
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 12
    t = ws.cell(2, 2, "GetGreenr Stock Bulk Update — Match Review")
    t.font = Font(bold=True, size=15, color=NAVY)
    for c, h in [(2, "Category"), (3, "Count")]:
        cell = ws.cell(4, c, h)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    for i, (name, cnt) in enumerate(counts, start=5):
        a = ws.cell(i, 2, name); b = ws.cell(i, 3, cnt)
        a.border = BORDER; b.border = BORDER
        b.alignment = Alignment(horizontal="center")
    return ws


# --------------------------------------------------------------------------- #
# Build
# --------------------------------------------------------------------------- #
def g(row, k):
    v = row.get(k, "")
    return "" if (v is None or (isinstance(v, float) and pd.isna(v))) else v


def main():
    out = M.build()
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    # ---------- Locked Matches ---------- #
    lk_headers = ["#", "GetGreenr Sheet", "GetGreenr SKU ID", "GetGreenr Model", "Storage",
                  "Connectivity", "GetGreenr Colour", "Condition Column",
                  "LOCKED Masterlist ID(s)", "ML Category", "ML Model(s)|Color",
                  "ML Available Qty", "Target Qty", "# SKUs"]
    lk_bands = [NAVY] + [ORANGE] * 7 + [YELLOW] * 5 + [NAVY]
    lk_widths = [5, 14, 40, 24, 9, 12, 16, 16, 22, 12, 55, 14, 11, 7]
    lk_rows = []
    for i, (_, r) in enumerate(out["locked"].iterrows(), 1):
        lk_rows.append([i, category(r["gg_sku"]), r["gg_sku"], disp_model(r["gg_sku"]),
                        g(r, "capacity"), connectivity(r["gg_sku"]), disp_colour(r["gg_sku"]),
                        g(r, "gg_on_hand") if False else "Excellent",
                        g(r, "linked_StockTypeIDs"), g(r, "ml_cats"), g(r, "ml_pipe"),
                        g(r, "summed_qty"), g(r, "summed_qty"), g(r, "n_links")])

    # ---------- Match Review ---------- #
    mr_headers = ["#", "GetGreenr Sheet", "GetGreenr SKU ID", "GetGreenr Model", "Storage",
                  "Connectivity", "GetGreenr Colour", "Condition Column", "Current Qty",
                  "Match Status", "Suggested Masterlist ID", "ML Category", "ML Model|Color",
                  "Score", "Reviewer Decision", "Notes"]
    mr_bands = [NAVY] + [ORANGE] * 8 + [YELLOW] * 7
    mr_widths = [5, 14, 40, 24, 9, 12, 16, 16, 11, 30, 22, 12, 45, 7, 22, 24]
    rev = out["review"].copy()
    if not rev.empty:
        rev["brand"] = rev["gg_sku"].str.split().str[0]
        rev = rev.sort_values(["brand", "reason", "model"])
    mr_rows = []
    for i, (_, r) in enumerate(rev.iterrows(), 1):
        decision = "Skip / Delist" if g(r, "status") == "No Masterlist match" else ""
        mr_rows.append([i, category(r["gg_sku"]), r["gg_sku"], disp_model(r["gg_sku"]),
                        g(r, "capacity"), connectivity(r["gg_sku"]), disp_colour(r["gg_sku"]),
                        "Excellent", g(r, "gg_on_hand"),
                        g(r, "status"), g(r, "candidate_StockTypeIDs"), g(r, "ml_cats"),
                        g(r, "ml_pipe"), g(r, "score"), decision, ""])

    # ---------- New Masterlist SKUs (reverse review) ---------- #
    nm_headers = ["#", "Masterlist Stock Type ID", "Category", "Brand", "Model", "Color",
                  "Available Qty", "Routed Condition", "Suggested GetGreenr SKU ID",
                  "Suggested Model", "Suggested Storage", "Suggested Colour",
                  "Match %", "Link to GetGreenr SKU ID", "Reviewer Decision", "Notes"]
    nm_bands = [NAVY] + [ORANGE] * 7 + [YELLOW] * 8
    nm_widths = [5, 20, 10, 12, 34, 16, 12, 16, 26, 22, 12, 16, 8, 24, 20, 20]
    nm_rows = []
    for i, (_, r) in enumerate(out["not_yet"].iterrows(), 1):
        cats = str(r.get("StockTypeIDs", ""))
        nm_rows.append([i, cats.split(",")[0].strip() if cats else "", "", r["brand"],
                        g(r, "raw_model"), g(r, "raw_color"), g(r, "qty"), "",
                        "", "", "", "", "", "", "", ""])

    # ---------- Not Selling in GetGreenr (manual) ---------- #
    ns_headers = ["#", "Masterlist Stock Type ID", "Category", "Brand", "Model", "Color",
                  "Available Qty", "Reason"]
    ns_bands = [NAVY] + [ORANGE] * 7
    ns_widths = [5, 20, 10, 12, 34, 16, 12, 26]

    # ---------- Not on GetGreenr Yet ---------- #
    ny_headers = ["#", "Masterlist Stock Type ID", "Category", "Brand", "Model", "Color",
                  "Available Qty", "Routed Condition"]
    ny_bands = [NAVY] + [ORANGE] * 7
    ny_widths = [5, 20, 10, 12, 34, 16, 12, 16]

    # ---------- Summary + write ---------- #
    counts = [
        ("Locked Matches (sync stock)", len(lk_rows)),
        ("New Masterlist SKUs (to add / review)", len(nm_rows)),
        ("Match Review (GetGreenr SKU needs check)", len(mr_rows)),
        ("Not Selling in GetGreenr", 0),
        ("Not on GetGreenr Yet", 0),
    ]
    nm_decisions = ["Linked", "Not on GetGreenr Yet", "Not Selling in GetGreenr"]
    mr_decisions = ["Skip / Delist", "Link manually", "Set to 0", "Keep as-is"]

    summary_sheet(wb, counts)
    write_sheet(wb, "Locked Matches", lk_headers, lk_bands, lk_rows, lk_widths)
    write_sheet(wb, "New Masterlist SKUs", nm_headers, nm_bands, nm_rows, nm_widths,
                dropdowns=[(15, nm_decisions)])
    write_sheet(wb, "Match Review", mr_headers, mr_bands, mr_rows, mr_widths,
                dropdowns=[(15, mr_decisions)])
    write_sheet(wb, "Not Selling in GetGreenr", ns_headers, ns_bands, [], ns_widths)
    write_sheet(wb, "Not on GetGreenr Yet", ny_headers, ny_bands, [], ny_widths)

    wb.save(OUT)
    print("Wrote", OUT)
    print("Sheets:", wb.sheetnames)
    print("Counts -> locked:%d review:%d new_ml:%d" %
          (len(lk_rows), len(mr_rows), len(nm_rows)))


if __name__ == "__main__":
    main()
